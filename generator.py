import os
import re
import urllib.request
import datetime
import json
from PIL import Image, ImageDraw, ImageFont
import openpyxl

try:
    from drive_sync import GoogleDriveSync
except ImportError:
    GoogleDriveSync = None

# Default fonts to download from Google Fonts if not locally present
FONT_URLS = {
    "Montserrat-Regular.ttf": "https://github.com/JulietaUla/Montserrat/raw/master/fonts/ttf/Montserrat-Regular.ttf",
    "Montserrat-Medium.ttf": "https://github.com/JulietaUla/Montserrat/raw/master/fonts/ttf/Montserrat-Medium.ttf",
    "Montserrat-SemiBold.ttf": "https://github.com/JulietaUla/Montserrat/raw/master/fonts/ttf/Montserrat-SemiBold.ttf",
    "Montserrat-Bold.ttf": "https://github.com/JulietaUla/Montserrat/raw/master/fonts/ttf/Montserrat-Bold.ttf",
    "PlayfairDisplay-Bold.ttf": "https://github.com/google/fonts/raw/main/ofl/playfairdisplay/PlayfairDisplay%5Bwght%5D.ttf"
}


def sanitize_filename(filename):
    """Sanitizes a string to make it safe for file names."""
    s = re.sub(r'(?u)[^-\w ]', '', filename)
    s = re.sub(r'\s+', ' ', s).strip()
    return s if s else "certificate"

def parse_members_string(members_str):
    """
    Parses a string formatted like:
    "Member 1: Full Name: Arya Jain, College/University: Parul ... | Member 2: Full Name: ERICK ATHNAS, ..."
    Returns a list of dictionaries, where each dictionary represents a member's parsed info.
    """
    if not members_str or not isinstance(members_str, str):
        return []
        
    members = []
    # Split by the pipe '|' character
    blocks = members_str.split('|')
    
    for block in blocks:
        block = block.strip()
        if not block:
            continue
            
        # We want to extract key-value pairs. 
        # A block typically looks like: "Member 1: Full Name: Arya Jain, College/University: Parul ..."
        # Let's clean the prefix (like "Member 1: ")
        if block.lower().startswith("member"):
            parts = block.split(':', 1)
            if len(parts) > 1:
                block = parts[1].strip()
                
        # Now block is: "Full Name: Arya Jain, College/University: Parul institute of technology, ..."
        # Common keys to extract:
        keys = ["Full Name", "College/University", "Current Course/Degree", "Email Address", "Mobile Number"]
        
        positions = []
        for k in keys:
            # Case-insensitive search for Key:
            idx = block.lower().find(k.lower() + ":")
            if idx != -1:
                positions.append((idx, k))
        
        # Sort positions
        positions.sort()
        
        member_info = {}
        for i in range(len(positions)):
            start_idx, key = positions[i]
            # The value starts after "Key:"
            val_start = start_idx + len(key) + 1
            # The value ends at the start of the next key, or the end of the block
            val_end = positions[i+1][0] if i + 1 < len(positions) else len(block)
            
            val = block[val_start:val_end].strip()
            # Strip trailing comma or spaces
            val = val.rstrip(',').strip()
            member_info[key] = val
            
        if member_info.get("Full Name"):
            members.append(member_info)
            
    return members

def download_fonts(fonts_dir="fonts"):
    """Downloads necessary fonts from Google Fonts if they are missing."""
    os.makedirs(fonts_dir, exist_ok=True)
    
    # Create an unverified SSL context to bypass certificate verification issues (common on Mac)
    context = None
    try:
        import ssl
        context = ssl._create_unverified_context()
    except Exception:
        pass
        
    for font_name, url in FONT_URLS.items():
        font_path = os.path.join(fonts_dir, font_name)
        if not os.path.exists(font_path):
            print(f"Downloading font: {font_name}...")
            try:
                # Use a standard user-agent to avoid HTTP blocks
                req = urllib.request.Request(
                    url, 
                    headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
                )
                
                open_kwargs = {"timeout": 15}
                if context is not None:
                    open_kwargs["context"] = context
                    
                with urllib.request.urlopen(req, **open_kwargs) as response, open(font_path, 'wb') as out_file:
                    out_file.write(response.read())
                print(f"Successfully downloaded {font_name}")
            except Exception as e:
                print(f"Warning: Failed to download {font_name} ({e}). Falling back to system fonts.")


def find_font_path(font_name_or_path, fonts_dir="fonts"):
    """
    Attempts to locate the given font path.
    1. Checks if it's a valid direct path.
    2. Checks in the local fonts directory.
    3. Checks standard OS-specific directories.
    4. Returns None if not found.
    """
    if os.path.exists(font_name_or_path):
        return font_name_or_path
        
    local_path = os.path.join(fonts_dir, font_name_or_path)
    if os.path.exists(local_path):
        return local_path

    # Check for direct file in fonts directory
    basename = os.path.basename(font_name_or_path)
    local_basename_path = os.path.join(fonts_dir, basename)
    if os.path.exists(local_basename_path):
        return local_basename_path

    # Standard system font search paths
    system_paths = []
    import platform
    current_os = platform.system().lower()
    
    if current_os == "darwin":  # macOS
        system_paths = [
            "/Library/Fonts",
            "/System/Library/Fonts",
            "/System/Library/Fonts/Supplemental",
            os.path.expanduser("~/Library/Fonts")
        ]
    elif current_os == "windows":
        system_paths = [os.path.join(os.environ.get("SystemRoot", "C:\\Windows"), "Fonts")]
    else:  # Linux / Unix
        system_paths = [
            "/usr/share/fonts",
            "/usr/share/fonts/truetype",
            "/usr/local/share/fonts",
            os.path.expanduser("~/.fonts")
        ]
        
    for sys_dir in system_paths:
        if os.path.exists(sys_dir):
            # Try exact match, case insensitive, or recursive check
            for root, _, files in os.walk(sys_dir):
                for f in files:
                    if f.lower() == basename.lower():
                        return os.path.join(root, f)
                        
    return None

def get_font(font_name_or_path, font_size, fonts_dir="fonts"):
    """Loads a font at the specified size with fallbacks."""
    font_path = find_font_path(font_name_or_path, fonts_dir)
    if font_path:
        try:
            return ImageFont.truetype(font_path, font_size)
        except Exception as e:
            print(f"Error loading truetype font {font_path}: {e}")
            
    # Try a default standard font name depending on OS
    fallback_fonts = ["Arial.ttf", "Helvetica.ttf", "Times New Roman.ttf", "Courier New.ttf"]
    for fallback in fallback_fonts:
        path = find_font_path(fallback, fonts_dir)
        if path:
            try:
                return ImageFont.truetype(path, font_size)
            except Exception:
                continue
                
    # Absolute fallback to Pillow's basic default font
    print("Warning: Loading default Pillow font (size control not supported).")
    return ImageFont.load_default()

class CertificateGenerator:
    def __init__(self, config=None):
        self.config = config or {}
        self.fonts_dir = self.config.get("fonts_dir", "fonts")
        # Ensure fonts are ready
        download_fonts(self.fonts_dir)

    def read_excel_data(self, excel_path):
        """Reads recipient data from either an Excel file (.xlsx) or CSV file (.csv)."""
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"File not found at: {excel_path}")
            
        print(f"Reading data from {excel_path}...")
        records = []
        
        if excel_path.lower().endswith('.csv'):
            import csv
            try:
                # Try reading with utf-8 first, fallback to latin-1
                try:
                    with open(excel_path, mode='r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        headers = [h.strip() for h in next(reader)]
                        for row_idx, row in enumerate(reader, start=2):
                            if not any(row_val for row_val in row):
                                continue
                            record = {}
                            for col_idx, val in enumerate(row):
                                if col_idx < len(headers):
                                    header = headers[col_idx]
                                    if header:
                                        record[header] = val.strip()
                            records.append((row_idx, record))
                except UnicodeDecodeError:
                    with open(excel_path, mode='r', encoding='latin-1') as f:
                        reader = csv.reader(f)
                        headers = [h.strip() for h in next(reader)]
                        for row_idx, row in enumerate(reader, start=2):
                            if not any(row_val for row_val in row):
                                continue
                            record = {}
                            for col_idx, val in enumerate(row):
                                if col_idx < len(headers):
                                    header = headers[col_idx]
                                    if header:
                                        record[header] = val.strip()
                            records.append((row_idx, record))
            except Exception as e:
                raise ValueError(f"Failed to parse CSV file: {str(e)}")
        else:
            # Excel file (.xlsx)
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            # Get headers from the first row
            headers = [cell.value for cell in sheet[1]]
            if not headers or all(h is None for h in headers):
                raise ValueError("Excel sheet is empty or has no header row.")
                
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row_val is not None for row_val in row):
                    continue  # Skip empty rows
                    
                record = {}
                for col_idx, val in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        if header is not None:
                            # Clean values, formatting dates nicely
                            if isinstance(val, (datetime.datetime, datetime.date)):
                                val = val.strftime("%B %d, %Y")
                            elif val is None:
                                val = ""
                            else:
                                val = str(val).strip()
                            record[header] = val
                records.append((row_idx, record))
                
        print(f"Loaded {len(records)} data rows.")
        return records

    def generate_single(self, template_image, record, fields_config):
        """Overlays text fields onto a single copy of the template image."""
        # Work on a copy of the template
        img = template_image.copy()
        draw = ImageDraw.Draw(img)
        
        for field_name, f_cfg in fields_config.items():
            # Extract configuration properties
            x = f_cfg.get("x", img.width // 2)
            y = f_cfg.get("y", img.height // 2)
            font_name = f_cfg.get("font", "Montserrat-Regular.ttf")
            font_size = f_cfg.get("font_size", 32)
            color = f_cfg.get("color", "#000000")
            align = f_cfg.get("align", "center").lower()
            
            # Map user alignment config to Pillow text anchor
            # 'anchor' controls the coordinate origin of the drawn text
            # 'm' = middle, 'l' = left, 'r' = right, 't' = top, 'm' = middle (vertical)
            anchor_map = {
                "left": "lt",
                "center": "mt",
                "right": "rt"
            }
            anchor = anchor_map.get(align, "mt")
            
            # Extract text to render (can use python template format)
            text_template = f_cfg.get("text", "{" + field_name + "}")
            try:
                text_to_draw = text_template.format(**record)
            except KeyError:
                # If template references a key not in the record, try direct field name
                text_to_draw = record.get(field_name, text_template)
            
            # Render the text
            font = get_font(font_name, font_size, self.fonts_dir)
            draw.text((x, y), text_to_draw, fill=color, font=font, anchor=anchor)
            
        return img

    def run(self, template_path, excel_path, output_dir="output", preview_only=False, progress_callback=None):
        """Generates certificates for all records in the Excel sheet."""
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template image not found at: {template_path}")
            
        # Load template
        print(f"Loading template image: {template_path}...")
        template_img = Image.open(template_path)
        
        # Load data
        records = self.read_excel_data(excel_path)
        if not records:
            print("No records found to generate certificates.")
            return
            
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Initialize Google Drive Sync client
        drive_sync = None
        drive_links = {}
        links_json_path = os.path.join(output_dir, "drive_links.json")
        
        # Load existing drive links if file exists
        if os.path.exists(links_json_path):
            try:
                with open(links_json_path, 'r') as f:
                    drive_links = json.load(f)
            except Exception:
                pass
                
        if GoogleDriveSync:
            sync_client = GoogleDriveSync()
            if sync_client.is_connected():
                drive_sync = sync_client
                # Ensure the main dynamic Certificates folder exists
                main_folder_name = self.config.get("main_folder_name", "Certificates")
                main_folder_id = drive_sync.find_or_create_folder(main_folder_name)
                drive_sync.main_folder_id = main_folder_id
            else:
                print("Google Drive Sync: Credentials not found or auth failed. Running in Local-Only mode.")
        else:
            print("Google Drive Sync: Google client libraries not installed. Running in Local-Only mode.")

        fields_config = self.config.get("fields", {})
        if not fields_config:
            print("Warning: No fields config defined in configuration. Using basic default mapping.")
            # Default fallback fields mapping if config is empty
            fields_config = {
                "Name": {"x": template_img.width // 2, "y": template_img.height // 2 - 50, "font_size": 48, "font": "Montserrat-Bold.ttf", "align": "center"},
                "Role": {"x": template_img.width // 2, "y": template_img.height // 2 + 50, "font_size": 28, "font": "Montserrat-Medium.ttf", "align": "center"},
                "Event": {"x": template_img.width // 2, "y": template_img.height // 2 + 120, "font_size": 32, "font": "Montserrat-SemiBold.ttf", "align": "center"},
                "Date": {"x": template_img.width // 2, "y": template_img.height // 2 + 200, "font_size": 24, "font": "Montserrat-Regular.ttf", "align": "center"}
            }
            
        if preview_only:
            records = records[:1]
            print("Running in PREVIEW mode: Generating certificate for the first row only...")

        # Determine scanned columns based on configuration checkboxes (dynamic list or fallback)
        cols_to_scan = self.config.get("scanned_columns", None)
        
        if cols_to_scan is None:
            # Fallback to legacy config
            include_leader = self.config.get("include_leader", True)
            include_members = self.config.get("include_members", True)
            include_members_leader = self.config.get("include_members_leader", True)
            
            cols_to_scan = []
            if include_members:
                cols_to_scan.append("Add members ( 2 to 4 )")
            if include_members_leader:
                cols_to_scan.append("Add members ( 2 to 4 )  including team leader ")
            if include_leader:
                cols_to_scan.append("Leader")
            cols_to_scan.append("Members")

        # 1. PRE-CALCULATE UNIQUE RECIPIENTS PER ROW (for accurate totals)
        total_people_count = 0
        people_per_row = []
        
        for row_idx, record in records:
            row_people = []
            seen_in_row = set()
            
            for col in cols_to_scan:
                if col in record and record[col]:
                    val_str = str(record[col]).strip()
                    if not val_str:
                        continue
                    if "|" in val_str or "Full Name:" in val_str:
                        for m in parse_members_string(val_str):
                            m_name = m.get("Full Name", "").strip()
                            if m_name and m_name.lower() not in seen_in_row:
                                row_people.append({"name": m_name, "info": m})
                                seen_in_row.add(m_name.lower())
                    else:
                        for part in val_str.replace('\n', ',').split(','):
                            name = part.strip()
                            if name and name.lower() not in seen_in_row:
                                row_people.append({"name": name, "info": None})
                                seen_in_row.add(name.lower())
                                
            if not row_people:
                for val in record.values():
                    if val and len(str(val)) > 2 and not str(val).isdigit() and not any(k in str(val).lower() for k in ["http", "@"]):
                        row_people.append({"name": str(val), "info": None})
                        break
                        
            people_per_row.append(row_people)
            total_people_count += len(row_people)

        # 2. BATCH GENERATION WITH CALLBACK CHECKS
        generated_count = 0
        is_cancelled = False
        
        for idx, (row_idx, record) in enumerate(records):
            if is_cancelled:
                break
                
            row_people = people_per_row[idx]
            
            # Determine team subfolder structure
            team_id = record.get("Team ID") or record.get("TeamID") or record.get("team_id") or ""
            team_name = record.get("Team Name") or record.get("TeamName") or record.get("team_name") or ""
            
            if team_id or team_name:
                raw_folder_name = f"{team_id} - {team_name}".strip(" -")
                folder_name = sanitize_filename(raw_folder_name)
                if not folder_name:
                    folder_name = "General"
            else:
                folder_name = ""
                raw_folder_name = ""
                
            # Create local team subfolder if applicable
            if folder_name:
                team_local_dir = os.path.join(output_dir, folder_name)
                os.makedirs(team_local_dir, exist_ok=True)
                
            for person in row_people:
                person_name = person["name"]
                member_info = person["info"]
                
                # Check for cancellation before rendering
                if progress_callback:
                    if not progress_callback(generated_count, total_people_count, person_name):
                        print("Generation cancelled by request.")
                        is_cancelled = True
                        break
                        
                safe_name = sanitize_filename(person_name)
                output_filename = f"certificate_{safe_name}.png"
                
                if folder_name:
                    relative_file_path = f"{folder_name}/{output_filename}"
                    output_path = os.path.join(team_local_dir, output_filename)
                else:
                    relative_file_path = output_filename
                    output_path = os.path.join(output_dir, output_filename)
                
                # Construct person-specific record
                person_record = record.copy()
                person_record["Name"] = person_name
                person_record["Leader"] = person_name
                
                # Overwrite with member specific details if available
                if member_info:
                    if "College/University" in member_info:
                        val = member_info["College/University"]
                        person_record["College_University Name"] = val
                        person_record["College/University"] = val
                    if "Current Course/Degree" in member_info:
                        val = member_info["Current Course/Degree"]
                        person_record["Current Course_Degree"] = val
                        person_record["Current Course/Degree"] = val
                    if "Email Address" in member_info:
                        val = member_info["Email Address"]
                        person_record["Email"] = val
                        person_record["Email Address"] = val
                    if "Mobile Number" in member_info:
                        val = member_info["Mobile Number"]
                        person_record["Leader Mobile Number"] = val
                        person_record["Mobile Number"] = val
                        
                print(f"Generating certificate for: {person_name} (Team: {raw_folder_name or 'None'})...")
                try:
                    cert_img = self.generate_single(template_img, person_record, fields_config)
                    # Keep quality high (save as PNG or high-quality JPG depending on output extension)
                    cert_img.save(output_path, dpi=(300, 300))
                    generated_count += 1
                    
                    # Google Drive cloud sync upload
                    if drive_sync and not preview_only:
                        # 1. Create or find the subfolder on Google Drive
                        drive_folder_name = folder_name if folder_name else "General"
                        drive_subfolder_id = drive_sync.find_or_create_folder(drive_folder_name, drive_sync.main_folder_id)
                        # 2. Upload the file to Google Drive
                        drive_link = drive_sync.upload_file(output_path, drive_subfolder_id)
                        if drive_link:
                            drive_links[relative_file_path] = drive_link
                except Exception as e:
                    print(f"Error generating certificate for {person_name}: {e}")
                    
        # Trigger final callback on success
        if progress_callback and not is_cancelled:
            progress_callback(generated_count, total_people_count, "Completed")
            
        # Save drive links mapping file if links were created
        if drive_links:
            try:
                with open(links_json_path, 'w') as f:
                    json.dump(drive_links, f, indent=4)
                print(f"Saved {len(drive_links)} Google Drive link mapping(s) to: {links_json_path}")
            except Exception as e:
                print(f"Error saving drive links mapping file: {e}")
                
        print(f"\nDone! Successfully generated {generated_count} certificate(s) in: {os.path.abspath(output_dir)}")


def generate_sample_assets(project_dir):
    """Generates sample template image, Excel data, and config.json in the project folder."""
    print("Creating sample assets for quick start...")
    os.makedirs(project_dir, exist_ok=True)
    
    # 1. Download fonts first so we can use them to draw on the sample template
    fonts_dir = os.path.join(project_dir, "fonts")
    download_fonts(fonts_dir)
    
    # 2. Draw a beautiful, premium sample template image (1920x1080)
    template_path = os.path.join(project_dir, "sample_template.png")
    
    # Base colors
    bg_color = (250, 248, 245)      # Soft cream
    navy_color = (15, 23, 42)       # Dark slate navy
    gold_color = (180, 140, 70)     # Muted gold
    grey_color = (100, 110, 120)    # Soft slate grey
    
    img = Image.new("RGB", (1920, 1080), bg_color)
    draw = ImageDraw.Draw(img)
    
    # Outer navy border
    draw.rectangle([(40, 40), (1880, 1040)], outline=navy_color, width=12)
    # Inner gold border
    draw.rectangle([(55, 55), (1865, 1025)], outline=gold_color, width=4)
    
    # Corner geometric details
    # Top-Left corner accents
    draw.line([(40, 90), (90, 40)], fill=gold_color, width=3)
    draw.line([(40, 105), (105, 40)], fill=navy_color, width=2)
    # Top-Right corner accents
    draw.line([(1880, 90), (1830, 40)], fill=gold_color, width=3)
    draw.line([(1880, 105), (1815, 40)], fill=navy_color, width=2)
    # Bottom-Left corner accents
    draw.line([(40, 990), (90, 1040)], fill=gold_color, width=3)
    draw.line([(40, 975), (105, 1040)], fill=navy_color, width=2)
    # Bottom-Right corner accents
    draw.line([(1880, 990), (1830, 1040)], fill=gold_color, width=3)
    draw.line([(1880, 975), (1815, 1040)], fill=navy_color, width=2)

    # Static Texts on Template (with fallbacks if font fails)
    font_title = get_font("PlayfairDisplay-Bold.ttf", 64, fonts_dir)
    font_subtitle = get_font("Montserrat-Medium.ttf", 20, fonts_dir)
    font_desc_static = get_font("Montserrat-Regular.ttf", 20, fonts_dir)
    font_sig_label = get_font("Montserrat-Regular.ttf", 16, fonts_dir)
    font_signature = get_font("PlayfairDisplay-Bold.ttf", 28, fonts_dir) # Simulate a handwritten-looking font if possible

    # Header Title
    draw.text((960, 210), "CERTIFICATE OF RECOGNITION", fill=navy_color, font=font_title, anchor="mt")
    
    # Subtitle
    draw.text((960, 340), "THIS IS PROUDLY PRESENTED TO", fill=grey_color, font=font_subtitle, anchor="mt")
    
    # Dynamic Name will go at y=430 (represented in config.json)
    
    # Static description line 1
    draw.text((960, 570), "for their outstanding contribution and dedication as a", fill=grey_color, font=font_desc_static, anchor="mt")
    
    # Dynamic Role will go at y=635 (represented in config.json)
    
    # Static description line 2
    draw.text((960, 715), "during the event", fill=grey_color, font=font_desc_static, anchor="mt")
    
    # Dynamic Event Name will go at y=775 (represented in config.json)
    
    # Bottom details (Date and Signature placeholders)
    # Date Label static text (Dynamic date will go below this)
    draw.text((450, 880), "DATE OF ISSUE", fill=grey_color, font=font_subtitle, anchor="ma")
    draw.line([(300, 940), (600, 940)], fill=gold_color, width=2)
    # Dynamic Date will go at x=450, y=950 (represented in config.json)
    
    # Signature line
    draw.text((1470, 880), "AUTHORIZED SIGNATURE", fill=grey_color, font=font_subtitle, anchor="ma")
    draw.line([(1320, 940), (1620, 940)], fill=gold_color, width=2)
    # Draw a simulated beautiful signature
    draw.text((1470, 905), "Sophia Sterling", fill=navy_color, font=font_signature, anchor="ma")

    img.save(template_path, dpi=(300, 300))
    print(f"Sample template image saved to: {template_path}")
    
    # 3. Create a beautiful config.json
    import json
    config_path = os.path.join(project_dir, "config.json")
    config_data = {
        "template_path": "sample_template.png",
        "excel_path": "sample_data.xlsx",
        "output_dir": "output",
        "fonts_dir": "fonts",
        "fields": {
            "Name": {
                "x": 960,
                "y": 420,
                "font_size": 55,
                "color": "#0F172A",
                "font": "PlayfairDisplay-Bold.ttf",
                "align": "center"
            },
            "Role": {
                "x": 960,
                "y": 630,
                "font_size": 32,
                "color": "#B48C46",
                "font": "Montserrat-SemiBold.ttf",
                "align": "center"
            },
            "Event": {
                "x": 960,
                "y": 770,
                "font_size": 36,
                "color": "#0F172A",
                "font": "Montserrat-Bold.ttf",
                "align": "center"
            },
            "Date": {
                "x": 450,
                "y": 950,
                "font_size": 20,
                "color": "#475569",
                "font": "Montserrat-Regular.ttf",
                "align": "center"
            }
        }
    }
    
    with open(config_path, "w") as f:
        json.dump(config_data, f, indent=4)
    print(f"Sample config.json saved to: {config_path}")

    # 4. Create sample Excel file (sample_data.xlsx)
    excel_path = os.path.join(project_dir, "sample_data.xlsx")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Certificates"
    
    headers = ["Name", "Event", "Date", "Role", "Team ID", "Team Name"]
    sheet.append(headers)
    
    rows = [
        ["Jane Doe", "Global Dev Summit 2026", "2026-06-15", "Keynote Speaker", "T-101", "Team Alpha"],
        ["John Smith", "Artificial Intelligence Workshop", "2026-06-18", "Participant", "T-101", "Team Alpha"],
        ["Alex Rivera", "Cloud Computing Bootcamp", "2026-07-01", "Lead Instructor", "T-102", "Team Beta"],
        ["Sarah Jenkins", "Cybersecurity Hackathon", "2026-07-10", "First Place Winner", "T-103", "Team Gamma"]
    ]
    for r in rows:
        sheet.append(r)
        
    wb.save(excel_path)
    print(f"Sample Excel sheet saved to: {excel_path}")
