import os
import json
import zipfile
import threading
from io import BytesIO
from flask import Flask, request, jsonify, send_from_directory, send_file, render_template
from werkzeug.utils import secure_filename
from generator import CertificateGenerator, sanitize_filename, download_fonts, parse_members_string

app = Flask(__name__, template_folder="templates")

# Configure directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
FONTS_DIR = os.path.join(BASE_DIR, "fonts")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(FONTS_DIR, exist_ok=True)

# Write environment credentials to local files for cloud compatibility on startup
sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
if sa_json:
    try:
        with open(os.path.join(BASE_DIR, "service_account.json"), "w") as f:
            f.write(sa_json)
        print("Successfully wrote service_account.json from environment variable.")
    except Exception as e:
        print(f"Error writing service_account.json: {e}")

token_json = os.environ.get("GOOGLE_OAUTH_TOKEN_JSON")
if token_json:
    try:
        with open(os.path.join(BASE_DIR, "token.json"), "w") as f:
            f.write(token_json)
        print("Successfully wrote token.json from environment variable.")
    except Exception as e:
        print(f"Error writing token.json: {e}")

# Helper to load config
def load_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "template_path": "sample_template.png",
        "excel_path": "sample_data.xlsx",
        "output_dir": "output",
        "fonts_dir": "fonts",
        "fields": {}
    }

# Helper to save config
def save_config(config):
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=4)

@app.route("/health")
def health():
    return {"status": "ok"}

@app.route("/")
def index():
    # Make sure sample assets exist if someone visits page
    config = load_config()
    from generator import generate_sample_assets
    if not os.path.exists(os.path.join(BASE_DIR, config.get("template_path", "sample_template.png"))):
        try:
            generate_sample_assets(BASE_DIR)
        except Exception as e:
            print(f"Error generating sample assets: {e}")
    return render_template("index.html")

@app.route("/api/config", methods=["GET", "POST"])
def manage_config():
    if request.method == "POST":
        new_config = request.json
        current_config = load_config()
        # Merge values
        current_config.update(new_config)
        save_config(current_config)
        return jsonify({"status": "success", "config": current_config})
    
    return jsonify(load_config())

@app.route("/api/fonts", methods=["GET"])
def get_fonts():
    # Download standard fonts if they are missing
    try:
        download_fonts(FONTS_DIR)
    except Exception:
        pass
    
    fonts = []
    if os.path.exists(FONTS_DIR):
        fonts = [f for f in os.listdir(FONTS_DIR) if f.lower().endswith(('.ttf', '.otf'))]
    
    # Also add standard system fallback representations
    fonts.extend(["Arial.ttf", "Helvetica.ttf", "Times New Roman.ttf"])
    # Remove duplicates while preserving order
    unique_fonts = []
    for f in fonts:
        if f not in unique_fonts:
            unique_fonts.append(f)
            
    return jsonify(unique_fonts)

@app.route("/api/data", methods=["GET"])
def get_excel_data():
    config = load_config()
    excel_path = os.path.join(BASE_DIR, config.get("excel_path", ""))
    
    if not os.path.exists(excel_path):
        return jsonify({"error": "Excel file not found. Please upload one."}), 404
        
    try:
        generator = CertificateGenerator(config=config)
        records = generator.read_excel_data(excel_path)
        # Strip index and return records list
        raw_list = [rec[1] for rec in records]
        
        # Get headers/columns from config or excel directly
        headers = []
        if raw_list:
            headers = list(raw_list[0].keys())
            
        return jsonify({"headers": headers, "records": raw_list})
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel: {str(e)}"}), 500

@app.route("/api/upload", methods=["POST"])
def upload_files():
    config = load_config()
    
    if "template" in request.files:
        file = request.files["template"]
        if file.filename != "":
            filename = secure_filename(file.filename)
            save_path = os.path.join(BASE_DIR, filename)
            file.save(save_path)
            config["template_path"] = filename
            
    if "excel" in request.files:
        file = request.files["excel"]
        if file.filename != "":
            filename = secure_filename(file.filename)
            save_path = os.path.join(BASE_DIR, filename)
            file.save(save_path)
            config["excel_path"] = filename
            
    save_config(config)
    return jsonify({"status": "success", "config": config})

# Thread-safe global generation state
generation_state = {
    "is_running": False,
    "current": 0,
    "total": 0,
    "current_name": "",
    "status": "idle",  # "idle", "running", "completed", "cancelled", "failed"
    "error": None,
    "cancel_requested": False,
    "files": []
}
generation_lock = threading.Lock()

def make_progress_callback():
    def callback(current, total, current_name):
        with generation_lock:
            generation_state["current"] = current
            generation_state["total"] = total
            generation_state["current_name"] = current_name
            if generation_state["cancel_requested"]:
                generation_state["status"] = "cancelled"
                generation_state["is_running"] = False
                return False
            return True
    return callback

def run_generation_thread(config, template_path, excel_path, output_dir):
    global generation_state
    try:
        callback = make_progress_callback()
        generator = CertificateGenerator(config=config)
        generator.run(
            template_path=template_path,
            excel_path=excel_path,
            output_dir=output_dir,
            preview_only=False,
            progress_callback=callback
        )
        
        with generation_lock:
            if generation_state["status"] != "cancelled":
                generation_state["status"] = "completed"
                # Scan output directory for generated files
                generated_files = []
                if os.path.exists(output_dir):
                    for root, dirs, files in os.walk(output_dir):
                        for file in files:
                            if file.lower().endswith(('.png', '.jpg', '.jpeg')) and file != "preview.png" and not file.startswith('.'):
                                file_path = os.path.join(root, file)
                                rel_path = os.path.relpath(file_path, output_dir)
                                generated_files.append(rel_path)
                    generated_files.sort()
                generation_state["files"] = generated_files
    except Exception as e:
        import traceback
        traceback.print_exc()
        with generation_lock:
            generation_state["status"] = "failed"
            generation_state["error"] = str(e)
    finally:
        with generation_lock:
            generation_state["is_running"] = False

@app.route("/api/generate", methods=["POST"])
def generate_certificates():
    global generation_state
    config = load_config()
    preview_only = request.json.get("preview", False) if request.json else False
    
    template_path = os.path.join(BASE_DIR, config.get("template_path", ""))
    excel_path = os.path.join(BASE_DIR, config.get("excel_path", ""))
    output_dir = os.path.join(BASE_DIR, config.get("output_dir", "output"))
    
    if not os.path.exists(template_path):
        return jsonify({"error": "Template image not found. Please upload a template."}), 400
    if not os.path.exists(excel_path):
        return jsonify({"error": "Excel file not found. Please upload an Excel sheet."}), 400
        
    try:
        # Save any configuration updates sent with the generate request
        if request.json:
            if "fields" in request.json:
                config["fields"] = request.json["fields"]
            if "include_leader" in request.json:
                config["include_leader"] = request.json["include_leader"]
            if "include_members" in request.json:
                config["include_members"] = request.json["include_members"]
            if "include_members_leader" in request.json:
                config["include_members_leader"] = request.json["include_members_leader"]
            if "scanned_columns" in request.json:
                config["scanned_columns"] = request.json["scanned_columns"]
            save_config(config)
            
        if preview_only:
            generator = CertificateGenerator(config=config)
            generator.run(
                template_path=template_path,
                excel_path=excel_path,
                output_dir=output_dir,
                preview_only=True
            )
            return jsonify({
                "status": "success",
                "message": "Successfully generated preview certificate.",
                "files": ["certificate_preview.png"]
            })
            
        # Start background thread for full batch generation
        with generation_lock:
            if generation_state["is_running"]:
                return jsonify({"error": "A generation task is already running."}), 400
                
            generation_state.update({
                "is_running": True,
                "current": 0,
                "total": 0,
                "current_name": "Initializing...",
                "status": "running",
                "error": None,
                "cancel_requested": False,
                "files": []
            })
            
        thread = threading.Thread(
            target=run_generation_thread,
            args=(config, template_path, excel_path, output_dir)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({
            "status": "success",
            "message": "Certificate generation started in background."
        })
        
    except Exception as e:
        return jsonify({"error": f"Generation failed: {str(e)}"}), 500

@app.route("/api/generate/status", methods=["GET"])
def get_generation_status():
    with generation_lock:
        return jsonify(generation_state)

@app.route("/api/generate/stop", methods=["POST"])
def stop_generation():
    global generation_state
    with generation_lock:
        if generation_state["is_running"]:
            generation_state["cancel_requested"] = True
            generation_state["status"] = "cancelled"
            return jsonify({"status": "success", "message": "Cancellation request submitted."})
        return jsonify({"error": "No generation task is currently running."}), 400


@app.route("/api/preview", methods=["GET"])
def get_preview():
    config = load_config()
    template_path = os.path.join(BASE_DIR, config.get("template_path", ""))
    excel_path = os.path.join(BASE_DIR, config.get("excel_path", ""))
    
    if not os.path.exists(template_path) or not os.path.exists(excel_path):
        return jsonify({"error": "Missing template or Excel data"}), 400
        
    try:
        generator = CertificateGenerator(config=config)
        records = generator.read_excel_data(excel_path)
        if not records:
            return jsonify({"error": "No data rows in Excel"}), 400
            
        from PIL import Image
        template_img = Image.open(template_path)
        
        # Get first record
        _, first_record = records[0]
        
        fields_config = config.get("fields", {})
        preview_img = generator.generate_single(template_img, first_record, fields_config)
        
        # Save in output as preview.png
        preview_path = os.path.join(OUTPUT_DIR, "preview.png")
        preview_img.save(preview_path)
        
        return jsonify({"status": "success", "url": "/output/preview.png"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/output/<path:filename>")
def serve_output(filename):
    return send_from_directory(OUTPUT_DIR, filename)

@app.route("/<path:filename>")
def serve_root_files(filename):
    # Security check: Only serve image files from the base directory
    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
        return send_from_directory(BASE_DIR, filename)
    return "Not Found", 404


@app.route("/api/download-all", methods=["GET"])
def download_all():
    config = load_config()
    output_dir = os.path.join(BASE_DIR, config.get("output_dir", "output"))
    main_folder_name = config.get("main_folder_name", "Certificates")
    safe_folder_name = sanitize_filename(main_folder_name) or "Certificates"
    
    if not os.path.exists(output_dir):
        return jsonify({"error": "No generated files found"}), 404
        
    # Walk output directory recursively to find all generated certificates
    certificate_files = []
    for root, dirs, files in os.walk(output_dir):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg')) and file != "preview.png" and not file.startswith('.'):
                file_path = os.path.join(root, file)
                # Save relative path within output directory
                arcname = os.path.relpath(file_path, output_dir)
                certificate_files.append((file_path, arcname))
                
    if not certificate_files:
        return jsonify({"error": "No certificates generated yet"}), 404
        
    # Create zip file in memory
    memory_file = BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file_path, arcname in certificate_files:
            # Place files inside a main folder named safe_folder_name
            zip_file.write(file_path, os.path.join(safe_folder_name, arcname))
            
    memory_file.seek(0)
    
    return send_file(
        memory_file,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{safe_folder_name}.zip"
    )

@app.route("/api/download-excel-report", methods=["GET"])
def download_excel_report():
    config = load_config()
    excel_path = os.path.join(BASE_DIR, config.get("excel_path", ""))
    
    if not os.path.exists(excel_path):
        return jsonify({"error": "Original spreadsheet file not found"}), 404
        
    try:
        import datetime
        import openpyxl
        from openpyxl.styles import Font
        
        # Load data
        if excel_path.lower().endswith('.csv'):
            import csv
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Certificates"
            
            rows_list = []
            try:
                try:
                    with open(excel_path, mode='r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            rows_list.append(row)
                except UnicodeDecodeError:
                    with open(excel_path, mode='r', encoding='latin-1') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            rows_list.append(row)
            except Exception as e:
                return jsonify({"error": f"Failed to parse CSV file: {str(e)}"}), 500
                
            for r in rows_list:
                sheet.append(r)
        else:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
        # Get headers from row 1
        headers = [cell.value for cell in sheet[1]]
        if not headers or all(h is None for h in headers):
            return jsonify({"error": "Spreadsheet has no header row."}), 400
            
        # Load drive links if they exist
        drive_links = {}
        links_json_path = os.path.join(OUTPUT_DIR, "drive_links.json")
        if os.path.exists(links_json_path):
            try:
                with open(links_json_path, 'r') as f:
                    drive_links = json.load(f)
            except Exception:
                pass
                
        # Scan all rows to find unique people per row and calculate max people
        rows_data = []
        max_people = 0
        for row_idx in range(2, sheet.max_row + 1):
            row_dict = {}
            for col_idx, header in enumerate(headers, start=1):
                if header is not None:
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.strftime("%B %d, %Y")
                    elif val is None:
                        val = ""
                    else:
                        val = str(val).strip()
                    row_dict[header] = val
            
            # Parse members for this row
            members_list = []
            for col in ["Add members ( 2 to 4 )", "Add members ( 2 to 4 )  including team leader ", "Members"]:
                if col in row_dict and row_dict[col]:
                    val_str = row_dict[col]
                    if "|" in val_str or "Full Name:" in val_str:
                        members_list.extend(parse_members_string(val_str))
                    else:
                        for part in val_str.replace('\n', ',').split(','):
                            name = part.strip()
                            if name:
                                members_list.append({"Full Name": name})
                                
            leader_name = row_dict.get("Leader", row_dict.get("Name", "")).strip()
            
            # Unique people
            people = []
            seen_names_lower = set()
            if leader_name:
                people.append({"name": leader_name})
                seen_names_lower.add(leader_name.lower())
                
            for m in members_list:
                m_name = m.get("Full Name", "").strip()
                if m_name and m_name.lower() not in seen_names_lower:
                    people.append({"name": m_name})
                    seen_names_lower.add(m_name.lower())
                    
            if not people:
                for val in row_dict.values():
                    if val and len(val) > 2 and not val.isdigit():
                        people.append({"name": val})
                        break
                        
            max_people = max(max_people, len(people))
            rows_data.append((row_idx, row_dict, people))
            
        # Add columns dynamically
        start_col = len(headers) + 1
        for i in range(max_people):
            col_name = f"Certificate Link ({'Leader' if i == 0 else f'Member {i}'})"
            sheet.cell(row=1, column=start_col + i, value=col_name)
            
        # Fill cells
        for row_idx, row_dict, people in rows_data:
            team_id = row_dict.get("Team ID") or row_dict.get("TeamID") or row_dict.get("team_id") or ""
            team_name = row_dict.get("Team Name") or row_dict.get("TeamName") or row_dict.get("team_name") or ""
            if team_id or team_name:
                raw_folder_name = f"{team_id} - {team_name}".strip(" -")
                folder_name = sanitize_filename(raw_folder_name)
                if not folder_name:
                    folder_name = "General"
            else:
                folder_name = ""
                
            for i, person in enumerate(people):
                person_name = person["name"]
                safe_name = sanitize_filename(person_name)
                output_filename = f"certificate_{safe_name}.png"
                
                if folder_name:
                    relative_file_path = f"{folder_name}/{output_filename}"
                    output_path = os.path.join(OUTPUT_DIR, folder_name, output_filename)
                else:
                    relative_file_path = output_filename
                    output_path = os.path.join(OUTPUT_DIR, output_filename)
                    
                cell = sheet.cell(row=row_idx, column=start_col + i)
                
                # Check Drive link
                drive_url = drive_links.get(relative_file_path)
                if drive_url:
                    cell.value = f"Link ({person_name})"
                    cell.hyperlink = drive_url
                    cell.font = Font(color="0563C1", underline="single")
                elif os.path.exists(output_path):
                    url = f"{request.host_url}output/{relative_file_path}"
                    cell.value = f"Local ({person_name})"
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = "Not Generated"
                    cell.font = Font(color="808080", italic=True)
                    
        # Save to stream
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        main_folder_name = config.get("main_folder_name", "Certificates")
        safe_folder_name = sanitize_filename(main_folder_name) or "Certificates"
        
        return send_file(
            out_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{safe_folder_name}_report.xlsx"
        )
    except Exception as e:
        return jsonify({"error": f"Failed to generate Excel report: {str(e)}"}), 500

@app.route("/api/certificates", methods=["GET"])
def list_certificates():
    config = load_config()
    output_dir = os.path.join(BASE_DIR, config.get("output_dir", "output"))
    
    generated_files = []
    if os.path.exists(output_dir):
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                if file.lower().endswith(('.png', '.jpg', '.jpeg')) and file != "preview.png" and not file.startswith('.'):
                    file_path = os.path.join(root, file)
                    rel_path = os.path.relpath(file_path, output_dir)
                    generated_files.append(rel_path)
        generated_files.sort()
        
    drive_links = {}
    links_json_path = os.path.join(output_dir, "drive_links.json")
    if os.path.exists(links_json_path):
        try:
            with open(links_json_path, 'r') as f:
                drive_links = json.load(f)
        except Exception:
            pass
            
    return jsonify({
        "files": generated_files,
        "drive_links": drive_links
    })

@app.route("/api/certificates/delete-all", methods=["POST"])
def delete_all_certificates():
    config = load_config()
    output_dir = os.path.join(BASE_DIR, config.get("output_dir", "output"))
    
    if not os.path.exists(output_dir):
        return jsonify({"status": "success", "message": "No certificates found to delete."})
        
    try:
        import shutil
        for item in os.listdir(output_dir):
            item_path = os.path.join(output_dir, item)
            if item == "preview.png":
                continue
            try:
                if os.path.isdir(item_path):
                    shutil.rmtree(item_path)
                else:
                    os.remove(item_path)
            except Exception as e:
                print(f"Error deleting {item_path}: {e}")
                
        return jsonify({
            "status": "success",
            "message": "All generated certificates deleted successfully."
        })
    except Exception as e:
        return jsonify({"error": f"Failed to delete certificates: {str(e)}"}), 500

@app.route("/api/drive-status", methods=["GET", "POST", "DELETE"])
def drive_status():
    from drive_sync import GoogleDriveSync
    
    # Use get_json(silent=True) to avoid raising HTTP 415 when content-type is not JSON
    json_data = request.get_json(silent=True) or {}
    
    if request.method == "DELETE" or (request.method == "POST" and json_data.get("action") == "disconnect"):
        target = json_data.get("target") or "all"
        warning = None
        if os.environ.get("RENDER") or os.environ.get("HOST") == "0.0.0.0":
            warning = "Note: Since you are deployed on Render, if you configured Google Drive Sync using environment variables, you must delete or clear them from your Render dashboard to fully remove access."
            
        if target == "service_account":
            path = os.path.join(BASE_DIR, "service_account.json")
            if os.path.exists(path):
                try: os.remove(path)
                except: pass
        elif target == "oauth":
            for filename in ["credentials.json", "token.json"]:
                path = os.path.join(BASE_DIR, filename)
                if os.path.exists(path):
                    try: os.remove(path)
                    except: pass
        else:
            # Clear all
            for filename in ["service_account.json", "credentials.json", "token.json"]:
                path = os.path.join(BASE_DIR, filename)
                if os.path.exists(path):
                    try: os.remove(path)
                    except: pass
                    
        return jsonify({"status": "success", "connected": False, "warning": warning})
        
    if request.method == "POST":
        # Handle OAuth login request
        if json_data.get("action") == "oauth_login":
            if os.environ.get("RENDER") or os.environ.get("HOST") == "0.0.0.0":
                return jsonify({
                    "status": "error",
                    "error": "Interactive OAuth login is not supported on cloud hosting environments. Please paste the content of your token.json file directly into the GOOGLE_OAUTH_TOKEN_JSON environment variable in your Render dashboard."
                })
                
            token_path = os.path.join(BASE_DIR, "token.json")
            if os.path.exists(token_path):
                try: os.remove(token_path)
                except: pass
            sync_client = GoogleDriveSync(method="oauth", interactive=True)
            is_connected = sync_client.is_connected()
            return jsonify({"status": "success", "connected": is_connected})
            
        # Handle method selection update
        if "active_method" in json_data:
            new_method = json_data["active_method"] # "service_account", "oauth", or "none"
            config = load_config()
            config["active_drive_method"] = new_method
            save_config(config)
            return jsonify({"status": "success", "active_method": new_method})
            
        # Handle file uploads
        if "service_account" in request.files:
            file = request.files["service_account"]
            if file.filename != "":
                save_path = os.path.join(BASE_DIR, "service_account.json")
                file.save(save_path)
                
                # Make active
                config = load_config()
                config["active_drive_method"] = "service_account"
                save_config(config)
                
                sync_client = GoogleDriveSync(method="service_account", interactive=False)
                return jsonify({
                    "status": "success", 
                    "connected": sync_client.is_connected(), 
                    "type": "service_account"
                })
                
        if "credentials" in request.files:
            file = request.files["credentials"]
            if file.filename != "":
                save_path = os.path.join(BASE_DIR, "credentials.json")
                file.save(save_path)
                # Clear previous token
                token_path = os.path.join(BASE_DIR, "token.json")
                if os.path.exists(token_path):
                    try: os.remove(token_path)
                    except: pass
                    
                # Make active
                config = load_config()
                config["active_drive_method"] = "oauth"
                save_config(config)
                
                # Check connection without interactive flow on upload
                sync_client = GoogleDriveSync(method="oauth", interactive=False)
                return jsonify({
                    "status": "success", 
                    "connected": sync_client.is_connected(), 
                    "type": "oauth"
                })
                
    # GET status info
    config = load_config()
    active_method = config.get("active_drive_method")
    
    has_sa = os.path.exists(os.path.join(BASE_DIR, "service_account.json")) or bool(os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON"))
    has_oauth = os.path.exists(os.path.join(BASE_DIR, "credentials.json"))
    has_token = os.path.exists(os.path.join(BASE_DIR, "token.json")) or bool(os.environ.get("GOOGLE_OAUTH_TOKEN_JSON"))
    
    if not active_method:
        if has_sa:
            active_method = "service_account"
        elif has_oauth:
            active_method = "oauth"
        else:
            active_method = "none"
            
    # Check connections for both separately to display their individual status
    sa_client = GoogleDriveSync(method="service_account", interactive=False)
    sa_connected = sa_client.is_connected()
    
    oauth_client = GoogleDriveSync(method="oauth", interactive=False)
    oauth_connected = oauth_client.is_connected()
    
    # Determine overall connected state based on the active selection
    is_connected = False
    if active_method == "service_account":
        is_connected = sa_connected
    elif active_method == "oauth":
        is_connected = oauth_connected
        
    return jsonify({
        "connected": is_connected,
        "active_method": active_method,
        "service_account": {
            "has_file": has_sa,
            "connected": sa_connected
        },
        "oauth": {
            "has_file": has_oauth,
            "has_token": has_token,
            "connected": oauth_connected
        }
    })


if __name__ == "__main__":
    import os
    # Load configuration from environment variables for deployment flexibility
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", 5001))
    debug_mode = os.environ.get("FLASK_DEBUG", "false").lower() in ("true", "1", "t")
    
    print(f"Starting Certificate Generator Web UI...")
    print(f"Please open http://{host}:{port} in your browser.")
    app.run(host=host, port=port, debug=debug_mode)
